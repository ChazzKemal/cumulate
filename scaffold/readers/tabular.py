"""Spreadsheet loading that survives real-world files.

Real exports have merged header cells, a title row above the headers, dates stored
as text, blank spacer columns, duplicate column names, and totals rows at the bottom.
Handle it here once so no tool has to rediscover it.
"""
from __future__ import annotations

import re
from pathlib import Path

import pandas as pd

EXTENSIONS = (".xlsx", ".xls", ".xlsm", ".csv", ".tsv")


def list_sheets(path: str | Path) -> list[str]:
    return pd.ExcelFile(path).sheet_names


def profile(path: str | Path) -> dict:
    """Everything you need to describe a file to someone before touching it."""
    path = Path(path)
    if path.suffix.lower() in {".csv", ".txt"}:
        sheets = {"(csv)": load(path)}
    else:
        sheets = {s: load(path, sheet=s) for s in list_sheets(path)}

    out = {"file": path.name, "kind": "tabular", "warnings": [], "sheets": {}}
    for name, df in sheets.items():
        out["sheets"][name] = {
            "rows": len(df),
            "columns": list(df.columns),
            "dtypes": {c: str(t) for c, t in df.dtypes.items()},
            "null_counts": {c: int(n) for c, n in df.isna().sum().items() if n},
            "sample": df.head(5).to_dict("records"),
            "warnings": _warnings(df),
        }
    return out


def _warnings(df: pd.DataFrame) -> list[str]:
    w = []
    for col in df.columns:
        s = df[col]
        if s.isna().all():
            w.append(f"'{col}' is entirely empty")
            continue
        if s.dtype == object and _looks_like_dates(s):
            w.append(f"'{col}' looks like dates stored as text")
        if s.dtype == object and _looks_like_numbers(s):
            w.append(f"'{col}' looks like numbers stored as text")
        if pd.api.types.is_numeric_dtype(s) and (s < 0).any():
            w.append(f"'{col}' contains negative values ({int((s < 0).sum())} rows)")
    if df.duplicated().any():
        w.append(f"{int(df.duplicated().sum())} fully duplicate rows")
    if len(df) and df.iloc[-1].isna().sum() > len(df.columns) / 2:
        w.append("last row is mostly blank — possibly a totals row")
    return w


def _looks_like_dates(s: pd.Series) -> bool:
    sample = s.dropna().astype(str).head(20)
    if sample.empty:
        return False
    pat = re.compile(r"^\d{1,4}[-/.]\d{1,2}[-/.]\d{1,4}")
    return (sample.str.match(pat).mean()) > 0.8


def _looks_like_numbers(s: pd.Series) -> bool:
    sample = s.dropna().astype(str).str.strip().head(20)
    if sample.empty:
        return False
    cleaned = sample.str.replace(r"[,\s€$£%]", "", regex=True)
    return (cleaned.str.match(r"^-?\d+\.?\d*$").mean()) > 0.8


def load(path: str | Path, sheet: str | int = 0, header_row: int | None = None) -> pd.DataFrame:
    """Load a sheet, finding the real header row and cleaning up the usual mess."""
    path = Path(path)

    if path.suffix.lower() in {".csv", ".txt"}:
        raw = pd.read_csv(path, header=None, dtype=object)
    else:
        raw = pd.read_excel(path, sheet_name=sheet, header=None, dtype=object)

    if header_row is None:
        header_row = _find_header_row(raw)

    df = raw.iloc[header_row + 1:].copy()
    df.columns = _clean_headers(raw.iloc[header_row])
    df = df.reset_index(drop=True)

    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    df = df.loc[:, [c for c in df.columns if not str(c).startswith("_unnamed")]]

    for col in df.columns:
        df[col] = _coerce(df[col])

    return df.reset_index(drop=True)


def _find_header_row(raw: pd.DataFrame, scan: int = 15) -> int:
    """The header is the first dense row of mostly-text cells."""
    best, best_score = 0, -1.0
    for i in range(min(scan, len(raw))):
        row = raw.iloc[i]
        filled = row.notna().sum()
        if filled < 2:
            continue
        texty = sum(1 for v in row.dropna() if isinstance(v, str) and not _is_numeric_str(v))
        score = filled * 1.0 + texty * 1.5
        if score > best_score:
            best, best_score = i, score
    return best


def _is_numeric_str(v: str) -> bool:
    try:
        float(str(v).replace(",", "").strip())
        return True
    except ValueError:
        return False


def _clean_headers(row: pd.Series) -> list[str]:
    """Forward-fill merged headers, dedupe collisions, name the blanks."""
    out, last, seen = [], None, {}
    for i, v in enumerate(row):
        name = str(v).strip() if pd.notna(v) and str(v).strip() else None
        if name is None:
            name = last if last else f"_unnamed_{i}"
        else:
            last = name
        name = re.sub(r"\s+", " ", name)
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        out.append(name)
    return out


def _coerce(s: pd.Series) -> pd.Series:
    """Turn text-that-is-really-numbers-or-dates into the real thing."""
    if s.dtype != object:
        return s
    if _looks_like_numbers(s):
        cleaned = s.astype(str).str.replace(r"[,\s€$£]", "", regex=True)
        converted = pd.to_numeric(cleaned, errors="coerce")
        if converted.notna().sum() >= s.notna().sum() * 0.9:
            return converted
    if _looks_like_dates(s):
        converted = pd.to_datetime(s, errors="coerce", format="mixed", dayfirst=False)
        if converted.notna().sum() >= s.notna().sum() * 0.9:
            return converted
    return s


# --- formulas -------------------------------------------------------------
# Loading a spreadsheet gives you the numbers. The rules are in the formulas,
# and pandas throws those away. For a file somebody maintained for years, the
# formula bar is the documentation — so read it deliberately, not on every
# profile, because it means parsing the workbook a second time.

_FORMULA_EXT = {".xlsx", ".xlsm"}
_REF_DIGITS = re.compile(r"(?<=[A-Za-z$])(\d+)")


def _shape(formula: str) -> str:
    """Strip row numbers so a column of 5,000 copies collapses to one rule."""
    return _REF_DIGITS.sub("#", formula)


def formulas(path: str | Path, limit: int = 40) -> dict:
    """The distinct calculations in a workbook, most-used first.

    Returns {sheet: [{formula, example, count}]}. Empty for anything that is
    not a real .xlsx/.xlsm — .xls and .csv carry no formulas to read.
    """
    from collections import Counter

    path = Path(path)
    if path.suffix.lower() not in _FORMULA_EXT:
        return {}

    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        out = {}
        for ws in wb.worksheets:
            seen: Counter = Counter()
            example: dict[str, str] = {}
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        shape = _shape(v)
                        seen[shape] += 1
                        example.setdefault(shape, f"{cell.coordinate}: {v}")
            if seen:
                out[ws.title] = [
                    {"formula": s, "example": example[s], "count": n}
                    for s, n in seen.most_common(limit)
                ]
        return out
    finally:
        wb.close()
